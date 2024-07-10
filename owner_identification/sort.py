import openpyxl
from datetime import datetime

#グローバル変数
data_lst = []

#CSVの先頭部分を入れるためのflag
first_flag = True

#xlsxファイルを読み込んでリストに追加する
def read_xlsx(file_path):
    global first_flag  # グローバル変数を修正
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if first_flag and i == 0:
            data_lst.append(list(row))
            first_flag = False
        elif first_flag == False and i == 0:
            continue
        else:
            data_lst.append(list(row))

# Excelへデータを保存する関数
def save_xlsx(data_lst):
    #保存したいExcelファイルのパス
    excel_path = "/home/sora/basic_ex/outputs/jun_20-all.xlsx"
    #新しいワークブックを作成
    wb = openpyxl.Workbook()
    #アクティブなワークシートを取得
    ws = wb.active

    for row in data_lst:  # 修正
        print(f"row: {row}")
        ws.append(row)

    #Excel ファイルを保存
    wb.save(excel_path)

#データを日時でソートする関数
def sort_data_by_date(data_lst):
    #ヘッダーを除いた部分をソート
    sorted_data = [data_lst[0]] + sorted(data_lst[1:], key=lambda x: datetime.strptime(x[0], '%Y-%m-%dT%H:%M:%S.%f'))
    return sorted_data

if __name__ == "__main__":
    #結合したいxlsxファイルのリスト
    #'''
    xlsx_lst = [
        "/home/sora/basic_ex/outputs/jun_20-core-s1.xlsx",
        "/home/sora/basic_ex/outputs/jun_20-core-s2.xlsx",
        "/home/sora/basic_ex/outputs/jun_20-core-s3.xlsx"
    ]
    '''
    xlsx_lst = [
        "/home/sora/basic_ex/outputs/jun_20.xlsx",
        "/home/sora/basic_ex/outputs/jun_11-1.xlsx",
        "/home/sora/basic_ex/outputs/jun_11-2.xlsx",
    ]
    '''
    for xlsx_file in xlsx_lst:
        read_xlsx(xlsx_file)

    #データを日時でソート
    data_lst = sort_data_by_date(data_lst)

    #Excelファイルへ出力
    save_xlsx(data_lst)
